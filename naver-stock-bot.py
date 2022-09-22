# 1. 라이브러리 호출
import requests
from bs4 import BeautifulSoup as bs
import openpyxl
import datetime
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import chromedriver_autoinstaller
import os
import time
from selenium.webdriver.common.by import By
from emailSMTP import send
from discordwebhook import Discord
import datetime

dt_now = datetime.datetime.now()

selectList = ["거래량", "시가총액(억)", "PER(배)", "자산총계(억)", "부채총계(억)", "PBR(배)"]


"""
 페이지 스킵 함수입니다.
  Args : 
    content (string) : 하단 페이지 이동 텍스트 
  Return : 
    {boolean} 
"""
def skipList(content):
  skipContentList = ["맨앞", "다음", "맨뒤", "이전"]

  if content in skipContentList:
    return False
  else:
    return True



"""
 현재 페이지 페이징 함수입니다.
  Args : 
    page (int) : 현재 페이지 
    driver (driver) : 현재 드라이버
    cssSelect (elements) : 페이지 요소's
"""
def pagingMove(page, driver, cssSelect):
  for pageListCount in cssSelect:
    if (page in (11, 21, 31)):
      driver.find_element(By.CSS_SELECTOR,"#contentarea > div.box_type_l > table.Nnavi > tbody > tr > td.pgR > a").click()
      break
    else:
      if (skipList(pageListCount.text.replace(" ", ""))):
        if (int(pageListCount.text) == page):
          pageListCount.click()
          time.sleep(2)
          break



"""
 상단 검색조건 찾는 함수 입니다.
  Args : 
    cssSelect (elements) : 상단 검색 요소's
 
"""
def topListReset(cssSelect):
  for list in cssSelect:
    tdList = list.find_elements(By.TAG_NAME,"td")

    for inputTag in tdList:
      try:
        if (inputTag.find_element(By.TAG_NAME,"input").get_attribute('checked')):
          inputTag.find_element(By.TAG_NAME,"input").click()
      except NoSuchElementException:
        print("No such")

"""
 검색조건 변경 함수 입니다.
  Args : 
    cssSelect (elements) : 검색 요소's
 
"""
def search(cssSelect):
  for list in cssSelect:
    tdList = list.find_elements(By.TAG_NAME,"td")
  for inputTag in tdList:
    try:
      if (inputTag.text in selectList):
        inputTag.find_element(By.TAG_NAME,"input").click()
    except NoSuchElementException:
      print("No such")

"""
 크롤링봇 함수 입니다.
 
"""

def crawlBot():
  marketType = {
    "KOSPI": "0",
    "KOSDAQ": "1",
  }

  # 코스피 상품 객체 리스트 
  rankKospiList = []
  
  #  코스닥 상품 객체 리스트 
  rankKosdaqList = []

  class product:
    def __init__(self, rank, name,currentPrice,fullTime,flucTuationRate,faceValue,cap,totalAssets,totalDebt,operatingProfit,per,pbr):
      self.rank = rank
      self.name = name
      self.currentPrice = currentPrice
      self.fullTime = fullTime
      self.flucTuationRate = flucTuationRate
      self.faceValue = faceValue
      self.cap = cap
      self.totalAssets = totalAssets
      self.totalDebt = totalDebt
      self.operatingProfit = operatingProfit
      self.per = per
      self.pbr = pbr

  # KOSPI + KOSDAQ
  for market, code in marketType.items():
    if (market == "KOSPI"):
      wb = openpyxl.Workbook()
      ws = wb.active
      ws.title = market
    else:
      wb = openpyxl.load_workbook(
        "C:/Users/welgram-Inwoo/Desktop/네이버_증권_크롤링/" + str(
          dt_now.date()) + ".xlsx")
      wb.create_sheet(market)
      ws = wb[market]

    # 컬럼 정의
    ws['A1'] = "순위"
    ws['B1'] = "종목명"
    ws['C1'] = "현재가"
    ws['D1'] = "전일비"
    ws['E1'] = "등락률"
    ws['F1'] = "액면가"
    ws['G1'] = "시가총액"
    ws['H1'] = "자산총계"
    ws['I1'] = "부채총계"
    ws['J1'] = "영업이익"
    ws['K1'] = "PER"
    ws['L1'] = "PBR"
    ws['M1'] = "시가총액/영업이익"
    ws['N1'] = "순자산"
    ws['O1'] = "시가총액/순자산"

    # 변동적인 칼럼값 정의
    objRankCount = 1

    # 드라이버 연결
    driver = webdriver.Chrome()

    # 웹사이트 이동
    driver.get(
      f"https://finance.naver.com/sise/sise_market_sum.nhn?sosok={code}&page=1")

    topList = driver.find_elements(By.CSS_SELECTOR,"#contentarea_left > div.box_type_m > form > div > div > table > tbody > tr")

    # topList 초기화
    topListReset(topList)

    # 검색조건 변경
    search(topList)

    # 검색조건 서치
    driver.find_element(By.CSS_SELECTOR,"#contentarea_left > div.box_type_m > form > div > div > div > a:nth-child(1) > img").click()

    time.sleep(3)

    for page in range(1, 2):

      # 페이징 이동
      pageList = driver.find_elements(By.CSS_SELECTOR,"#contentarea > div.box_type_l > table.Nnavi > tbody > tr > td")
      pagingMove(page, driver, pageList)

      # 2. 데이터 호출
      req = driver.page_source
      soup = bs(req, "lxml")

      # 3. 데이터 추출 (파싱) 단계
      stockContents = soup.select(
        "#contentarea > div.box_type_l > table.type_2 > tbody > tr")
      for stockContent in stockContents:
        try:
          
          objRank = stockContent.select_one("td.no").text  # 순위
          objName = stockContent.select_one("td:nth-child(2) >a").text  # 종목명            

          objCurrentPrice = stockContent.select_one(
            "td:nth-child(3)").text  # 현재가
          objFullTime = stockContent.select_one(
            "td:nth-child(4) > span").text  # 전일비
          objFluctuationRate = stockContent.select_one(
            "td:nth-child(5) > span").text  # 등락률
          objFaceValue = stockContent.select_one("td:nth-child(6)").text  # 액면가
          objCap = stockContent.select_one("td:nth-child(7)").text  # 시가총액
          objTotalAssets = stockContent.select_one(
            "td:nth-child(8)").text  # 자산총계
          objTotalDebt = stockContent.select_one("td:nth-child(9)").text  # 부채총계
          objOperatingProfit = stockContent.select_one(
            "td:nth-child(10)").text  # 영업이익
          objPer = stockContent.select_one("td:nth-child(11)").text  # PER
          objPbr = stockContent.select_one("td:nth-child(12)").text  # PBR

          # 1 ~ 3위만 추출 
          if(int(objRank) < 4) :
            # 코스피일 경우 
            if(market == "KOSPI") :

              rankKospi = product(objRank,objName,objCurrentPrice,objFullTime,objFluctuationRate,objFaceValue,objCap,objTotalAssets,objTotalDebt,objOperatingProfit,objPer,objPbr)
              rankKospiList.append(rankKospi)
            else :# 코스닥일 경우 

              rankKosdaq = product(objRank,objName,objCurrentPrice,objFullTime,objFluctuationRate,objFaceValue,objCap,objTotalAssets,objTotalDebt,objOperatingProfit,objPer,objPbr)
              rankKosdaqList.append(rankKosdaq)

          ws['A' + str(int(objRank) + objRankCount)] = objRank
          ws['B' + str(int(objRank) + objRankCount)] = objName
          ws['C' + str(int(objRank) + objRankCount)] = objCurrentPrice
          ws['D' + str(int(objRank) + objRankCount)] = objFullTime
          ws['E' + str(int(objRank) + objRankCount)] = objFluctuationRate
          ws['F' + str(int(objRank) + objRankCount)] = objFaceValue
          ws['G' + str(int(objRank) + objRankCount)] = objCap
          ws['H' + str(int(objRank) + objRankCount)] = objTotalAssets
          ws['I' + str(int(objRank) + objRankCount)] = objTotalDebt
          ws['J' + str(int(objRank) + objRankCount)] = objOperatingProfit
          ws['K' + str(int(objRank) + objRankCount)] = objPer
          ws['L' + str(int(objRank) + objRankCount)] = objPbr

          print(f"{market} : {objRank}등 : 크롤링중....")
        except AttributeError:
          continue
        except ZeroDivisionError:
          continue
    print(f"{market} : 크롤링완료")

    wb.save("C:/Users/welgram-Inwoo/Desktop/네이버_증권_크롤링/" + str(
      dt_now.date()) + ".xlsx")
    driver.close()

  
  sendDiscord(rankKospiList,rankKosdaqList)
  
  return True


"""
 크롬드라이버 버전 체크 함수입니다.
  Retruns:
    {boolean}
"""

def chkDriver():
  # Check if chrome driver is installed or not
  chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
  driver_path = f'./{chrome_ver}/chromedriver.exe'

  if os.path.exists(driver_path):
    print(f"최신크롬 버전입니다 : {driver_path}")
    return True
  else:
    print(f"최신크롬 버전을 다운받습니다 (ver: {chrome_ver})")
    chromedriver_autoinstaller.install(True)
    return False


"""
 디스코드에 알림 함수 
  Args : 
    rankKospiList (objecList) : 코스피 상품 객체 리스트
    rankKosdaqList (objectList) : 코스닥 상품 객체 리스트 
 
"""

def sendDiscord(rankKospiList,rankKosdaqList):

    discord = Discord(url="")

    discord.post(
        embeds= [
            {
                "author" :{
                    "name" : "naver-stock-bot" , 
                    "icon_url" : "https://picsum.photos/400/300"
                }  , 
                "title" : "오늘의 네이버 주식 현황 [ "+str(dt_now.date())+" ]" , 
                "color": 0,
            } , 


            { 
              "title" : "코스피 [ 1등 ] : " +rankKospiList[0].name , 
              "color" : 1127128,
              "fields": [
                    {"name": "현재가", "value": rankKospiList[0].currentPrice, "inline": True},
                    {"name": "전일비", "value": rankKospiList[0].fullTime, "inline": True},
                    {"name": "등락률", "value": rankKospiList[0].flucTuationRate, "inline": True},

                    {"name": "액면가", "value": rankKospiList[0].faceValue, "inline": True},
                    {"name": "시가총액", "value": rankKospiList[0].cap, "inline": True},
                    {"name": "자산총계", "value": rankKospiList[0].totalAssets, "inline": True},
                    {"name": "부채총계", "value": rankKospiList[0].totalDebt, "inline": True},
                    {"name": "영업이익", "value": rankKospiList[0].operatingProfit, "inline": True},
                    {"name": "PER", "value": rankKospiList[0].per, "inline": True},
                    {"name": "PBR", "value": rankKospiList[0].pbr, "inline": True},
                ]
            } , 


            { 
              "title" : "코스피 [ 2등 ] : " +rankKospiList[1].name , 
              "color" : 1127128,
              "fields": [
                    {"name": "현재가", "value": rankKospiList[1].currentPrice, "inline": True},
                    {"name": "전일비", "value": rankKospiList[1].fullTime, "inline": True},
                    {"name": "등락률", "value": rankKospiList[1].flucTuationRate, "inline": True},

                    {"name": "액면가", "value": rankKospiList[1].faceValue, "inline": True},
                    {"name": "시가총액", "value": rankKospiList[1].cap, "inline": True},
                    {"name": "자산총계", "value": rankKospiList[1].totalAssets, "inline": True},
                    {"name": "부채총계", "value": rankKospiList[1].totalDebt, "inline": True},
                    {"name": "영업이익", "value": rankKospiList[1].operatingProfit, "inline": True},
                    {"name": "PER", "value": rankKospiList[1].per, "inline": True},
                    {"name": "PBR", "value": rankKospiList[1].pbr, "inline": True},
                ]
            } , 


            {
                "title" : "코스피 [ 3등 ] : " +rankKospiList[2].name , 
                "color" : 1127128,
                "fields": [
                    {"name": "현재가", "value": rankKospiList[2].currentPrice, "inline": True},
                    {"name": "전일비", "value": rankKospiList[2].fullTime, "inline": True},
                    {"name": "등락률", "value": rankKospiList[2].flucTuationRate, "inline": True},

                    {"name": "액면가", "value": rankKospiList[2].faceValue, "inline": True},
                    {"name": "시가총액", "value": rankKospiList[2].cap, "inline": True},
                    {"name": "자산총계", "value": rankKospiList[2].totalAssets, "inline": True},
                    {"name": "부채총계", "value": rankKospiList[2].totalDebt, "inline": True},
                    {"name": "영업이익", "value": rankKospiList[2].operatingProfit, "inline": True},
                    {"name": "PER", "value": rankKospiList[2].per, "inline": True},
                    {"name": "PBR", "value": rankKospiList[2].pbr, "inline": True},
                ]
            } , 


            { 
              "title" : "코스닥 [ 1등 ] : " +rankKosdaqList[0].name , 
              "color" : 14177041,
              "fields": [
                    {"name": "현재가", "value": rankKosdaqList[0].currentPrice, "inline": True},
                    {"name": "전일비", "value": rankKosdaqList[0].fullTime, "inline": True},
                    {"name": "등락률", "value": rankKosdaqList[0].flucTuationRate, "inline": True},

                    {"name": "액면가", "value": rankKosdaqList[0].faceValue, "inline": True},
                    {"name": "시가총액", "value": rankKosdaqList[0].cap, "inline": True},
                    {"name": "자산총계", "value": rankKosdaqList[0].totalAssets, "inline": True},
                    {"name": "부채총계", "value": rankKospiList[0].totalDebt, "inline": True},
                    {"name": "영업이익", "value": rankKosdaqList[0].operatingProfit, "inline": True},
                    {"name": "PER", "value": rankKosdaqList[0].per, "inline": True},
                    {"name": "PBR", "value": rankKosdaqList[0].pbr, "inline": True},
                ]
            } , 


            { 
              "title" : "코스닥 [ 2등 ] : " +rankKosdaqList[1].name , 
              "color" : 14177041,
              "fields": [
                    {"name": "현재가", "value": rankKosdaqList[1].currentPrice, "inline": True},
                    {"name": "전일비", "value": rankKosdaqList[1].fullTime, "inline": True},
                    {"name": "등락률", "value": rankKosdaqList[1].flucTuationRate, "inline": True},

                    {"name": "액면가", "value": rankKosdaqList[1].faceValue, "inline": True},
                    {"name": "시가총액", "value": rankKosdaqList[1].cap, "inline": True},
                    {"name": "자산총계", "value": rankKosdaqList[1].totalAssets, "inline": True},
                    {"name": "부채총계", "value": rankKosdaqList[1].totalDebt, "inline": True},
                    {"name": "영업이익", "value": rankKosdaqList[1].operatingProfit, "inline": True},
                    {"name": "PER", "value": rankKosdaqList[1].per, "inline": True},
                    {"name": "PBR", "value": rankKosdaqList[1].pbr, "inline": True},
                ]
            } , 


            {
                "title" : "코스닥 [ 3등 ] : " +rankKosdaqList[2].name , 
                "color" : 14177041,
                "fields": [
                    {"name": "현재가", "value": rankKosdaqList[2].currentPrice, "inline": True},
                    {"name": "전일비", "value": rankKosdaqList[2].fullTime, "inline": True},
                    {"name": "등락률", "value": rankKosdaqList[2].flucTuationRate, "inline": True},

                    {"name": "액면가", "value": rankKosdaqList[2].faceValue, "inline": True},
                    {"name": "시가총액", "value": rankKosdaqList[2].cap, "inline": True},
                    {"name": "자산총계", "value": rankKosdaqList[2].totalAssets, "inline": True},
                    {"name": "부채총계", "value": rankKosdaqList[2].totalDebt, "inline": True},
                    {"name": "영업이익", "value": rankKosdaqList[2].operatingProfit, "inline": True},
                    {"name": "PER", "value": rankKosdaqList[2].per, "inline": True},
                    {"name": "PBR", "value": rankKosdaqList[2].pbr, "inline": True},
                ]
            } 
        ],
    )

    discord.post(
        file={
            "file1" : open("C:/Users/welgram-Inwoo/Desktop/네이버_증권_크롤링/"+str(dt_now.date())+".xlsx","rb"),
        }
    )


"""
 크롤링 실행 메인 함수입니다.
 
"""
def excuteCraw():
  if(chkDriver()):
    try:
      if (crawlBot()):
        try:
          email_result = send(reveiver_email, title, str(dt_now.date()))
          print("이메일발송 : " + str(email_result))
        except Exception as e:
          print(e)
    except Exception as e :
      print(e)

if __name__ == '__main__':
  reveiver_email = 'leekh916@hanmail.net'
  title = "[알림] 네이버 주식 크롤링 봇 정보 수집 완료 [" + str(dt_now.date()) + "]"
  try:
    excuteCraw()
  except Exception as e :
    print(e)
    print("한번만더 실행합니다")
    excuteCraw()



