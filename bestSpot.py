# 1. 라이브러리 호출
import requests
from bs4 import BeautifulSoup as bs
import openpyxl
from selenium import webdriver
import chromedriver_autoinstaller
import os
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import datetime

dt_now = datetime.datetime.now()

class Store:
    
    """
    검색조건에 따른 사업장 클래스
    """

    def __init__(self,name: str,tag: str,address: str):
        """
        - name : 사업장 이름
        - tag : 사업장 유형
        - address : 사업장 주소 
        """
        self.name = name
        self.tag = tag
        self.address = address

class Excel:
    """
    엑셀 클래스
    """

    @staticmethod
    def make_excel(sheet_title:str):
        """
        엑셀 생성
        - sheet_title : 시트 이름
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title

        # 엑셀 컬럼 정의
        ws['A1'] = "사업장_이름"
        ws['B1'] = "사업장_유형"
        ws['C1'] = "사업장_주소"



class StoreScraper:
    """
    사업장 크롤링 클래스
    """

    @staticmethod
    def get_stores(input_area: str, input_tag: str):
        """
        사업장 정보 가져오기

        - input_area : 사업장 지역
        - input_tag  : 사업장 태그
        """
        poolFlag = True
        store_product_arr = []
        totalLength = 0

        # 엑셀 시트 생성 후 엑셀 헤더 세팅
        # Excel.make_excel(input_area)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = input_area +" "+input_tag

        # 엑셀 컬럼 정의
        ws['A1'] = "사업장_이름"
        ws['B1'] = "사업장_유형"
        ws['C1'] = "사업장_주소"

        options = webdriver.ChromeOptions()

        # 드라이버 연결
        driver = webdriver.Chrome(options=options)

        # 웹사이트 이동 (검색 키워드)
        driver.get(f"https://map.naver.com/p/search/{input_area}+ {input_tag}")
        time.sleep(5)

        # 검색어 다시 입력
        search_clear_el = driver.find_element(By.CSS_SELECTOR,"#section_content > div > div.sc-iwm9f4.jCPpmH > div > div > button.btn_clear")
        search_clear_el.click()

        search_el = driver.find_element(By.CSS_SELECTOR,"#section_content > div > div.sc-iwm9f4.jCPpmH > div > div > div > input")
        search_el.send_keys(f"{input_area} {input_tag}")
        search_el.send_keys(Keys.ENTER)

        time.sleep(5)

        # iframe 처리 
        driver.switch_to.frame("searchIframe")

        # footer 처리 ( 화살표 )
        while poolFlag:
           
           # 보여지는 푸터의 총 리스트 개수  
           totalLength = int(driver.execute_script("let footerEl = document.querySelectorAll('#app-root > div > div.XUrfU > div.zRM9F'); return footerEl[0].childNodes.length-2; "))
           
           # 푸터의 마지막 리스트 확인 후 마지막 값
           footerLastValue =  int(driver.execute_script("let footerEl = document.querySelectorAll('#app-root > div > div.XUrfU > div.zRM9F'); return footerEl[0].childNodes[footerEl[0].childNodes.length-2].text;"))

           # 토탈 리스트가 5개 미만일 경우  
           if(totalLength <5):
              poolFlag = False

           # 토탈 리스트가 5개 이상일 경우 
           else :
              
              # 다음 리스트 호출
              driver.execute_script("let footerEl = document.querySelectorAll('#app-root > div > div.XUrfU > div.zRM9F'); footerEl[0].childNodes[footerEl[0].childNodes.length-1].click();")
              time.sleep(3)
              
              # 푸터의 마지막 요소 접근
              lastElCss = driver.execute_script("let footerEl = document.querySelectorAll('#app-root > div > div.XUrfU > div.zRM9F'); let lastEl = footerEl[0].childNodes[footerEl[0].childNodes.length-1]; let lastSvgEl = lastEl.querySelector('svg'); let footerCss = window.getComputedStyle(lastSvgEl); return footerCss.getPropertyValue('opacity') ")

              # 푸터의 마지막 값 확인이 되었을 경우
              if(lastElCss == '0.4'):
                 poolFlag = False
                 totalLength = footerLastValue

        # 푸터 초기 인덱스로 이동
        for x in range(footerLastValue):
           driver.execute_script("let footerEl = document.querySelectorAll('#app-root > div > div.XUrfU > div.zRM9F'); footerEl[0].childNodes[0].click();")
           time.sleep(3)

        print("=========================")
        print(f"총 리스트는 {totalLength}")
        print("푸터 첫번째 페이지로 이동")
        print("=========================")

        for listCount in range(totalLength):
          
          # 스크롤 맨끝까지 내렸을 경우
          booCnt = 0 

          # 첫번째 시작 페이지를 제외한 
          if(listCount != 0):
            driver.execute_script("let footerEl = document.querySelectorAll('#app-root > div > div.XUrfU > div.zRM9F'); footerEl[0].childNodes[footerEl[0].childNodes.length-1].click();")
            time.sleep(5)

          # 스크롤 맨 처음으로 초기화 
          driver.execute_script("document.querySelector(\"#_pcmap_list_scroll_container\").scrollTop = 0")

          # 맨 끝으로 내리기
          while booCnt < 10:
              driver.execute_script("document.querySelector(\"#_pcmap_list_scroll_container\").scrollTop += 1000")
              booCnt += 1
              time.sleep(2)

          # 데이터 호출
          req = driver.page_source
          soup = bs(req, "lxml")

          # 데이터 추출 (파싱) 단계

          # 검색 결과 리스트
          mainContents = soup.select("#_pcmap_list_scroll_container > ul > li")

          # 스크롤 맨 처음으로 초기화 
          driver.execute_script("document.querySelector(\"#_pcmap_list_scroll_container\").scrollTop = 0")

          # tr 형태가 아닌 , 테이블 형태 전체 셀 idx 기준으로 
          for idx in range(len(mainContents)-1):
                
                time.sleep(3)

                # 사업장 이름 
                store_name = mainContents[idx].select("div.qbGlu > div > a:nth-child(1) > div > div > span.place_bluelink.YwYLL")[0].text

                # 사업장 유형 
                store_tag = mainContents[idx].select_one("div.qbGlu > div > a:nth-child(1) > div > div > span.YzBgS").text

                # 사업장 주소 파악을 위한 클릭 이벤트
                driver.execute_script("let addressEl = document.querySelector('#_pcmap_list_scroll_container > ul > li:nth-child("+str(idx+1)+") > div.qbGlu > div > div > div > span > a > span.JXQZb');  addressEl.click();")

                # 이벤트에 따른 page_source 다시 불러오기
                req = driver.page_source
                soup = bs(req, "lxml")

                # 검색 결과 리스트
                mainContents = soup.select("#_pcmap_list_scroll_container > ul > li")

                # 사업장 주소
                store_address = mainContents[idx].select("div.qbGlu > div > div > div > div > div:nth-child(1)")[0].text.replace("도로명","").replace("복사","")

                # 검색한 주소를 포함하고 있지 않다면 다음 검색 결과로
                if(input_area not in store_address):
                   print(f"{input_area} 검색한 지역이 아닙니다.")
                   continue

                time.sleep(5)
                
                # 사업장 객체에 데이터 세팅 
                store_product = Store(store_name,store_tag,store_address)
                store_product_arr.append(store_product)

                ws['A' + str(idx+2)] = store_name
                ws['B' + str(idx+2)] = store_tag
                ws['C' + str(idx+2)] = store_address

                print(f"------- #{listCount+1} --------")
                print(f"------- {str(idx+1)} 번째 --------- \n사업장 이름 :: {store_name}\n사업장 유형 :: {store_tag} \n사업장 주소 :: {store_address}\n크롤링중....")

                driver.execute_script("document.querySelector(\"#_pcmap_list_scroll_container\").scrollTop += 170")
                
                print("스크롤 처리...")
                print("")

          print(f"# {listCount+1} :: 지역 :: {input_area} 유형 :: {input_tag} :: 크롤링완료")
          
          wb.save("C:/Users/dlsdn_q9bmeyr/Desktop/최고의_자리_찾기/" + 
                  str(dt_now.date()) +str(input_area)+"_"+str(input_tag)+"_"+str(listCount+1)+".xlsx")
          
        driver.close()
  

"""
 크롬드라이버 버전 체크 함수입니다.
  Retruns:
    {boolean}
"""

def chkDriver():
  # Check if chrome driver is installed or not
  chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
  driver_path = f'./{chrome_ver}/chromedriver.exe'

  if os.path.exists(driver_path):
    print(f"최신크롬 버전입니다 : {driver_path}")
    return True
  else:
    print(f"최신크롬 버전을 다운받습니다 (ver: {chrome_ver})")
    chromedriver_autoinstaller.install(True)
    return False

"""
 크롤링 실행 메인 함수입니다.
 
"""
def excuteCraw():
  if(chkDriver()):
    try:
      StoreScraper.get_stores("의왕","아파트")
    except Exception as e :
      print(e)

if __name__ == '__main__':

  excuteCraw()



