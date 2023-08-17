# 1. 라이브러리 호출
import requests
from bs4 import BeautifulSoup as bs
import openpyxl
from selenium import webdriver
import chromedriver_autoinstaller
import os
import time
from selenium.webdriver.common.by import By
import datetime

dt_now = datetime.datetime.now()

selectList = ["회차", "월상환금", "납입원금", "대출이자", "대출잔금"]


"""
 페이지 스킵 함수입니다.
  Args : 
    content (string) : 하단 페이지 이동 텍스트 
  Return : 
    {boolean} 
"""
def skipList(content):
  skipContentList = ["맨앞", "다음", "맨뒤", "이전"]

  if content in skipContentList:
    return False
  else:
    return True



"""
 현재 페이지 페이징 함수입니다.
  Args : 
    page (int) : 현재 페이지 
    driver (driver) : 현재 드라이버
    cssSelect (elements) : 페이지 요소's
"""
def pagingMove(page, driver, cssSelect):
  for pageListCount in cssSelect:
    if (page in (11, 21, 31)):
      driver.find_element(By.CSS_SELECTOR,"#contentarea > div.box_type_l > table.Nnavi > tbody > tr > td.pgR > a").click()
      break
    else:
      if (skipList(pageListCount.text.replace(" ", ""))):
        if (int(pageListCount.text) == page):
          pageListCount.click()
          time.sleep(2)
          break

"""
 크롤링봇 함수 입니다.
"""

def crawlBot():

  inputType = {
    "원금균등": "50년",
  }

  # 상환방법 ( 상환기간 ) ;
  repayMent = []

  # 대출 클래스
  class product:
    def __init__(self, term, monthlyPayment,principalPaid,loanInterest,loanBalance):
      self.term = term
      self.monthlyPayment = monthlyPayment
      self.principalPaid = principalPaid
      self.loanInterest = loanInterest
      self.loanBalance = loanBalance

  for repay, code in inputType.items():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = repay +"_"+ code

    # 컬럼 정의
    ws['A1'] = "회차"
    ws['B1'] = "월상환금"
    ws['C1'] = "납입원금"
    ws['D1'] = "대출이자"
    ws['E1'] = "대출잔금"

    # 변동적인 칼럼값 정의
    objRankCount = 1

    # 드라이버 연결
    driver = webdriver.Chrome()

    # 웹사이트 이동
    driver.get(
        f"https://ezb.co.kr/calculator/finances/loan-repayment")

    time.sleep(20)

    # 2. 데이터 호출
    req = driver.page_source
    soup = bs(req, "lxml")

    termEl = soup.select("body > main > div > div > div.caci-row.extension > div.extension-inner.type-cmn > div > div.extension-column.round.hidden > div")
    monthlyPaymentEl = soup.select("body > main > div > div > div.caci-row.extension > div.extension-inner.type-cmn > div > div.extension-column.repayment.hidden > div")
    principalPaidEl = soup.select("body > main > div > div > div.caci-row.extension > div.extension-inner.type-cmn > div > div.extension-column.principal.hidden > div")
    loanInterestEl = soup.select("body > main > div > div > div.caci-row.extension > div.extension-inner.type-cmn > div > div.extension-column.interest.hidden > div")
    loanBalanceEl = soup.select("body > main > div > div > div.caci-row.extension > div.extension-inner.type-cmn > div > div.extension-column.balance.hidden > div")
    
    # 3. 데이터 추출 (파싱) 단계
    headerContents = soup.select(
        "body > main > div > div > div.caci-row.extension > div.extension-inner.type-cmn > div > div.extension-column.round.hidden > div")
    
    # tr 형태가 아닌 , 테이블 형태 전체 셀 idx 기준으로 
    for idx in range(len(headerContents)-1):
        
        # 회차
        term = termEl[idx+1].text
        # 월상환금
        monthlyPayment = monthlyPaymentEl[idx+1].text
        # 납입원금
        startIdx = principalPaidEl[idx+1].text.find("대출잔금")
        
        principalPaid = principalPaidEl[idx+1].text[0:startIdx].replace(" ","")
        # 대출이자
        loanInterest = loanInterestEl[idx+1].text
        # 대출잔금
        loanBalance = loanBalanceEl[idx+1].text

        payProduct = product(term, monthlyPayment,principalPaid,loanInterest,loanBalance)
        repayMent.append(payProduct)

        ws['A' + str(int(idx+2))] = term
        ws['B' + str(int(idx+2))] = monthlyPayment
        ws['C' + str(int(idx+2))] = principalPaid
        ws['D' + str(int(idx+2))] = loanInterest
        ws['E' + str(int(idx+2))] = loanBalance

        print(f"{term} 회차 : 크롤링중....")

    print(f"{repay} | {code} : 크롤링완료")
    
    wb.save("C:/Users/dlsdn_q9bmeyr/Desktop/특례보금자리론_금리비교/" + str(
        dt_now.date()) +str(repay+"_"+code)+".xlsx")
    driver.close()

  return True


"""
 크롬드라이버 버전 체크 함수입니다.
  Retruns:
    {boolean}
"""

def chkDriver():
  # Check if chrome driver is installed or not
  chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
  driver_path = f'./{chrome_ver}/chromedriver.exe'

  if os.path.exists(driver_path):
    print(f"최신크롬 버전입니다 : {driver_path}")
    return True
  else:
    print(f"최신크롬 버전을 다운받습니다 (ver: {chrome_ver})")
    chromedriver_autoinstaller.install(True)
    return False

"""
 크롤링 실행 메인 함수입니다.
 
"""
def excuteCraw():
  if(chkDriver()):
    try:
      crawlBot()
    except Exception as e :
      print(e)

if __name__ == '__main__':

  excuteCraw()



