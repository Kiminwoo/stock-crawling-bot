from openpyxl import load_workbook
import urllib.request
import json

wb = load_workbook(filename='info_xlsx/지역별_태권도_정보.xlsx')

for loop in range(len(wb.sheetnames)):

    ws = wb.worksheets[loop]

    temp = []
    addresses = []

    for i in range(2,ws.max_row+1):
        temp.append([])
        for cell in ws[i]:
            temp[-1].append(cell.value)

    for i in range(0, ws.max_row -1):
        addresses.append(temp[i][2])

    count = 2 

    for address in addresses:
        encoding_address = urllib.parse.quote_plus(address)
        url = f'https://naveropenapi.apigw.ntruss.com/map-geocode/v2/geocode?query={encoding_address}'
        request = urllib.request.Request(url)
        request.add_header("X-NCP-APIGW-API-KEY-ID","u9em6i295p")
        request.add_header("X-NCP-APIGW-API-KEY","hMCxycT8CcAvGDQgLbnniiZoYuNFtHAjqAKhcBO9")

        response = urllib.request.urlopen(request)
        response_code = response.getcode()

        if response_code == 200:
            try:
                response_body = response.read()
                data = json.loads(response_body)
                ws.cell(row=count, column=4).value = data['addresses'][0]['y']
                ws.cell(row=count, column=5).value = data['addresses'][0]['x']

                print(ws.cell(row=count,column=3).value)

                count += 1
            except:
                ws.cell(row=count , column=4).value = 'error'
                ws.cell(row=count , column=5).value = 'error'
                count += 1
                print(f'error : {count}')

wb.save(filename='info_xlsx/지역별_태권도_정보.xlsx')