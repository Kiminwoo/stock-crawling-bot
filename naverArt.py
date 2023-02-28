# 1. 라이브러리 호출
import requests
from bs4 import BeautifulSoup as bs
import openpyxl
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import chromedriver_autoinstaller
import os
import time
from selenium.webdriver.common.by import By
from emailSMTP import send
from discord_bot import sendDiscord
import datetime

dt_now = datetime.datetime.now()

selectList = ["거래량", "시가총액(억)", "PER(배)", "자산총계(억)", "부채총계(억)", "PBR(배)"]


"""
 페이지 스킵 함수입니다.
  Args : 
    content (string) : 하단 페이지 이동 텍스트 
  Return : 
    {boolean} 
"""
def skipList(content):
  skipContentList = ["맨앞", "다음", "맨뒤", "이전"]

  if content in skipContentList:
    return False
  else:
    return True



"""
 현재 페이지 페이징 함수입니다.
  Args : 
    page (int) : 현재 페이지 
    driver (driver) : 현재 드라이버
    cssSelect (elements) : 페이지 요소's
"""
def pagingMove(page, driver, cssSelect):
  for pageListCount in cssSelect:
    if (page in (11, 21, 31)):
      driver.find_element(By.CSS_SELECTOR,"#contentarea > div.box_type_l > table.Nnavi > tbody > tr > td.pgR > a").click()
      break
    else:
      if (skipList(pageListCount.text.replace(" ", ""))):
        if (int(pageListCount.text) == page):
          pageListCount.click()
          time.sleep(2)
          break

"""
 크롤링봇 함수 입니다.
"""

def crawlBot():

  inputType = {
    "TheStarryNight": "0",
  }

  # artInfo ( 명화 정보 list ) ;
  artList = []

  # 명화 클래스
  class product:
    def __init__(self, author, authorBirthday,artTitle,artBirthday,artContent):
      self.author = author
      self.authorBirthDay = authorBirthday
      self.artTitle = artTitle
      self.artBirthday = artBirthday
      self.artContent = artContent

  for artName, code in inputType.items():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "명화"

    # 컬럼 정의
    ws['A1'] = "작가이름"
    ws['B1'] = "작가출생년일"
    ws['C1'] = "작품이름"
    ws['D1'] = "작품만든년도"
    ws['E1'] = "작품내용"

    # 변동적인 칼럼값 정의
    objRankCount = 1

    # 드라이버 연결
    driver = webdriver.Chrome()

    # 웹사이트 이동
    driver.get(
        f"https://terms.naver.com/search.naver?query={artName}&searchType=&dicType=&subject=")

    # 검색조건 서치
    driver.find_element(By.CSS_SELECTOR,"#content > div.search_result_area > ul > li:nth-child(1) > div.thumb_area > div > a > span").click()

    time.sleep(5)

    # 2. 데이터 호출
    req = driver.page_source
    soup = bs(req, "lxml")

    # 3. 데이터 추출 (파싱) 단계
    artHeaderContents = soup.select(
        "#size_ct > div.cont_area > div > div.att_type > div > div > table > tbody > tr")

    for idx in range(len(artHeaderContents)):

      try:

        # if(idx)

        author = artHeaderContents.select_one("tr:nth-child(1) > td > a").text  # 작가
        authorBirthday = artHeaderContents.select_one("td > b:nth-child(5)").text  # 출생년도
        artBirthday = artHeaderContents.select_one("tr:nth-child(2) > td > a").text  # 작품 출시일
        artTitle = artName # 작품 이름
        artContent = soup.select("#size_ct > div.cont_area > div > div.cont_txt").text # 작품 내용

        artProduct = product(author, authorBirthday,artTitle,artBirthday,artContent)
        artList.append(artProduct)

        ws['A' + str(int(objRankCount))] = author
        ws['B' + str(int(objRankCount))] = authorBirthday
        ws['C' + str(int(objRankCount))] = artBirthday
        ws['D' + str(int(objRankCount))] = artTitle
        ws['E' + str(int(objRankCount))] = artContent

        print(f"{artName} : 크롤링중....")
    except AttributeError:
      continue
    except ZeroDivisionError:
      continue

    print(f"{artName} : 크롤링완료")

    wb.save("C:/Users/welgram-Inwoo/Desktop/명화/" + str(
        dt_now.date()) + ".xlsx")
    driver.close()

  return True


"""
 크롬드라이버 버전 체크 함수입니다.
  Retruns:
    {boolean}
"""

def chkDriver():
  # Check if chrome driver is installed or not
  chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
  driver_path = f'./{chrome_ver}/chromedriver.exe'

  if os.path.exists(driver_path):
    print(f"최신크롬 버전입니다 : {driver_path}")
    return True
  else:
    print(f"최신크롬 버전을 다운받습니다 (ver: {chrome_ver})")
    chromedriver_autoinstaller.install(True)
    return False

"""
 크롤링 실행 메인 함수입니다.
 
"""
def excuteCraw():
  if(chkDriver()):
    try:
      crawlBot()
    except Exception as e :
      print(e)

if __name__ == '__main__':

  excuteCraw()



